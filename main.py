import xlrd
import re
import openpyxl
from tkinter import Tk
from tkinter.filedialog import askopenfilename, asksaveasfilename, askdirectory
import time
import datetime
import os
import shutil
from colorama import init, Fore, Style
import traceback

init(autoreset=True)

allIntDZ = 0
allIntCalcDZ = 0

DoubleKZ = []
extKZ = []
extKZZuordnung = {}

zonePrices = {}
hourPrices = {}

KZOrders = {}
KZOrdersCalc = {}
extOrders = {}
extOrdersCalc = {}

edvBook = {}

dz = 0.99

Excemptions = {}
mindermenge = 0
sackTonne = 0

fehler = []
notCalculated = 0
extNotCalculated = 0
header = [""]*14
minus = 0

edvbookPath = ""
saveintern = ""
abrFilePath = ""
savespotPath = ""
extbookPath = ""

date = ""

allowedZones = []

def main():

    global edvBook
    global edvbookPath
    global abrFilePath

    loadConfig()

    edvBook = xlrd.open_workbook(edvbookPath)

    Tk().withdraw()

    print(Fore.CYAN + Style.BRIGHT + "Drücken Sie Enter, um die Fahrten-Datei auszuwählen")
    input()

    abrFile = askopenfilename(initialdir=abrFilePath)
    abrBook = xlrd.open_workbook(abrFile)
    abrSheet = abrBook.sheet_by_index(0)

    #TZ eingabe
    readTZ()
    createFolders()

    #Einlesen der Auftragsdatei
    print(Fore.GREEN + "Lese Fahrten...")
    readAbrSheet(abrSheet)
    time.sleep(0.5)
    # Preise laden
    print(Fore.GREEN + "Lade Preise...")
    getZonePreise()
    getStundenPreise()
    time.sleep(0.5)

    #Einlesen der Frächteraufstellung
    print(Fore.GREEN + "Lese Subfrächterkennzeichen...")
    extractSubLicensePLates()
    time.sleep(0.5)

    #Handle interne Sachen
    print(Fore.GREEN + "Interne Buchungen berechnen...")
    calculateInterns()
    time.sleep(0.5)
    print(Fore.GREEN + "Interne Buchungen erstellen...")
    printKFZ(KZOrdersCalc, False)
    time.sleep(0.5)

    #Aunsahmen
    print(Fore.GREEN + "Lese Ausnahmen für Subfrächterberechnung...")
    readExcemptionos()
    time.sleep(0.5)

    #Handle externe Sachen
    print(Fore.GREEN + "Extrahiere Subfrächterfahrten...")
    extractSubFahrten()
    time.sleep(0.5)
    print(Fore.GREEN + "Rechne Subfrächtergutschriften...")
    calculateSubcontractors()
    time.sleep(0.5)
    print(Fore.GREEN + "Erstelle Subfrächtergutschriften...\n\n")
    printKFZ(extOrdersCalc, True)
    time.sleep(0.5)

    #Abschluss
    consoleOut()
    print(Fore.GREEN + "\n\nErstelle Minusfahrten.xlsx...")
    minusOut()
    time.sleep(0.5)

    print(Fore.GREEN + Style.BRIGHT + "\n\nProgramm beendet - Drücken Sie Enter oder das rote X")
    input()

def createFolders():
    global date
    today = datetime.date.today()
    first = today.replace(day=1)
    lastmonth = first - datetime.timedelta(days=1)
    date = lastmonth.strftime("%Y_%m")

    exdir = saveintern+"Subfrächter_"+date
    indir = saveintern+date

    if os.path.exists(exdir):
        shutil.rmtree(exdir)

    if os.path.exists(indir):
        shutil.rmtree(indir)

    os.mkdir(exdir)
    os.mkdir(indir)

def readTZ():

    global dz

    print(Fore.CYAN + Style.BRIGHT + "Geben Sie bitte den Treibstoffzuschlag in % ein! (Bei Subfrächtergutschriften wird automatisch +2% Aufschlag gerechnet!)")
    dp = input()
    dp = dp.replace("%", "")
    dp = dp.replace(",", ".")
    dp = float(dp)
    dp /= 100
    dp += 1
    dz = dp

def consoleOut():

    if allIntDZ == 0:
        print(Fore.RED + "Konnte den Treibstoffzuschlag nicht in der Liste finden. Wenn Sie die Differenz sehen möchten, geben Sie bitte Treibstoff im Feld Kennz. ein!\n")

    print( "TZ laut grüner Liste: ", Style.BRIGHT + str(allIntDZ))
    print("Errechneter TZ: ", Style.BRIGHT + str(round(allIntCalcDZ, 2)))
    print(Style.BRIGHT + "\tDifferenz: ", Fore.YELLOW + Style.BRIGHT + str(round(allIntDZ-allIntCalcDZ, 2)))
    print(Style.BRIGHT + "\nAusgelassene Datensätze - interne Verbuchung: ", Style.BRIGHT + str(notCalculated))
    for f in fehler:
        if f[2] == 0:
            print("\tZeile: ", Fore.YELLOW + str((f[0]+1)), "  -  " , Fore.YELLOW + header[f[1]])

    print(Style.BRIGHT + "\nAusgelassene Datensätze - Subfrächtergutschriften: ", Style.BRIGHT + str(extNotCalculated))

    for f in fehler:
        if f[2] == 2:
            print("\tZeile: ", Fore.YELLOW + str((f[0])), "  -  " , Fore.YELLOW + header[f[1]])

    print(Style.BRIGHT + "\nSumme aller Minus-Fahrten: ", Style.BRIGHT + Fore.RED + str(round(minus, 2)))

def minusOut():

    global savespotPath

    minusBook = openpyxl.Workbook()
    minusSheet = minusBook.active

    for h in range(14):
        minusSheet.cell(1, h+1).value = header[h]

    minusSheet.cell(1, 15).value = "Zeile (Grüne Liste)"
    minusSheet.cell(1, 13).value = "Eingang = Ger.Kosten + Maut + TZ"
    minusSheet.cell(1, 16).value = "Subfrächtergutschift"

    mapping = [1,2,3,4,5,6,7,9,10,11,12,13,14,None,15,8,None,16,None,None]

    r = 1
    for f in fehler:
        if f[2] == 3:
            r += 1
            c = 0
            for key, val in f[3].items():
                if c <=21:
                    if mapping[c] is not None:
                        minusSheet.cell(r, mapping[c]).value = val
                    c += 1

    widths = [
        ['A', 15],
        ['B', 10],
        ['C', 10],
        ['D', 18],
        ['E', 25],
        ['F', 35],
        ['G', 25],
        ['H', 10],
        ['I', 10],
        ['J', 9],
        ['K', 9],
        ['L', 10],
        ['M', 35],
        ['N', 15],
        ['O', 20],
        ['P', 20],
        ['Q', 20],
        ['R', 20],
        ['S', 20],
        ['T', 20],
        ['U', 20]
    ]

    for i in range(0, minusSheet.max_column, 1):
        minusSheet.column_dimensions[widths[i][0]].width = widths[i][1]

    print(Fore.CYAN + Style.BRIGHT + "\nGeben Sie bitte an, wo die Minusfahrten-Liste abgespeichert werden soll. Drücken Sie ENTER zum auswählen.")
    input()

    #TODO; return!
    saveSpot = asksaveasfilename(initialdir=savespotPath, defaultextension=".xlsx", initialfile="Minusfahrten_"+date)

    minusBook.save(saveSpot)

def readAbrSheet(srcSheet):
    global notCalculated
    global allIntDZ
    global header

    for c in range(1, srcSheet.ncols):
        header[c-1] = srcSheet.cell(9, c).value

    for c in range(srcSheet.ncols-4, srcSheet.ncols):
        header[c-1] = srcSheet.cell(8, c).value

    for i in range(10, srcSheet.nrows):

        success = True

        data = dict()
        data['geraet'] = srcSheet.cell(i, 1).value
        data['lfs_datum'] = srcSheet.cell(i, 2).value
        data['lfs_nr'] = srcSheet.cell(i, 3).value
        data['art_lfrnt'] = srcSheet.cell(i, 4).value
        data['art'] = srcSheet.cell(i, 5).value
        data['kunden'] = srcSheet.cell(i, 6).value
        data['baustelle'] = srcSheet.cell(i, 7).value
        data['zone'] = srcSheet.cell(i, 9).value
        data['einheit'] = srcSheet.cell(i, 10).value
        data['menge'] = srcSheet.cell(i, 11).value
        data['stunden'] = srcSheet.cell(i, 12).value
        data['ger_kosten'] = srcSheet.cell(i, 13).value
        data['mautk'] = srcSheet.cell(i, 14).value
        data['anmerkungen'] = ""
        data['zeile'] = i+1

        if isinstance(data['zone'], str):
            data['zone'] = data['zone'].upper()
        else:
            data['zone'] = str(int(data['zone']))

        kz = srcSheet.cell(i, 8).value

        doFehler = True
        if "Treibstoff" in kz:
            allIntDZ = data['ger_kosten']
            doFehler = False

        if "Summe" in data['geraet']:
            doFehler = False

        plates = extractLicensePlates(kz, i)

        if(len(plates) > 1) and doFehler:
            fehler.append([i, 7, 1])

        if(len(plates) == 0) and doFehler:
            success = False
            fehler.append([i, 7, 0])

        if data['ger_kosten'] is None or data['ger_kosten'] == '' or data['ger_kosten'] == 0 and doFehler:
            fehler.append([i, 12, 0])
            success = False

        if data['lfs_nr'] is None or data['lfs_nr'] == '' or data['lfs_nr'] == 0:
            fehler.append([i, 2, 0])
            success = False

        if not doFehler:
            success = False

        if(success):
            if plates[0] not in KZOrders:
                KZOrders[plates[0]] = []
            data['kz'] = plates[0]
            KZOrders[plates[0]].append(data)
        elif doFehler:
            notCalculated += 1

def extractLicensePlates(orPlate, i):
    entr = orPlate.replace('-','')
    entr = entr.replace('_4A','')

    if orPlate == '':
        entr = "KEIN"

    notpattern = re.compile('[a-z]')
    findpattern = re.compile('([A-Z]|[0-9]){3,9}')

    platesFound = []

    plates = entr.split('+')
    for plate in plates:
        parts = plate.split(',')
        for part in parts:
            nono = notpattern.search(part)
            if nono is None:
                yesyes = findpattern.search(part)
                if yesyes is not None and len(part) < 10:
                    platesFound.append(part)


    return platesFound;

def extractSubLicensePLates():

    global extbookPath

    extBook = xlrd.open_workbook(extbookPath)
    extSheet = extBook.sheet_by_index(0)

    notpattern = re.compile('[a-z]')
    findpattern = re.compile('([A-Z]|[0-9]){3,9}')

    curName = ""

    for col in range(0, extSheet.ncols):
        for row in range(0, extSheet.nrows):
            val = extSheet.cell(row, col).value

            if row == 0:
                curName = val

            val = val.replace(' ','')
            val = val.replace('_4A','')
            vals = val.split('-')
            for part in vals:
                nono = notpattern.search(part)
                if nono is None:
                    yesyes = findpattern.search(part)
                    if yesyes is not None and len(part) < 10:
                        if part not in extKZ and curName != "Reder":
                            extKZ.append(part)
                        if part not in extKZZuordnung:
                            extKZZuordnung[part] = curName


def extractSubFahrten():

    global extNotCalculated

    for kz in extKZ:
        if kz in KZOrders and extKZZuordnung[kz] != "Reder":
            orders = []

            for fahrt in KZOrders[kz]:
                success = True

                found = False
                for z in allowedZones:
                    if fahrt['zone'] == z:
                        found = True
                if fahrt['zone'] is None or fahrt['zone'] == '' or not found:
                    if fahrt['einheit'] != 'pau':
                        fehler.append([fahrt['zeile'], 8, 2])
                        extNotCalculated += 1
                        success = False

                if fahrt['einheit'] == '' or fahrt['einheit'] is None or fahrt['einheit'] == 'stk':
                    extNotCalculated += 1
                    fehler.append([fahrt['zeile'], 9, 2])
                    success = False

                if fahrt['menge'] == '' or fahrt['menge'] == 0 or fahrt['menge'] is None:
                    if fahrt['einheit'] != 'pau':
                        if fahrt['stunden'] == '' or fahrt['stunden'] == 0 or fahrt['stunden'] is None:
                            extNotCalculated += 1
                            fehler.append([fahrt['zeile'], 10, 2])
                            fehler.append([fahrt['zeile'], 11, 2])
                            success = False

                if fahrt['geraet'] == '' or fahrt['geraet'] == 0 or fahrt['geraet'] is None:
                    if fahrt['einheit'] != 'pau':
                        fehler.append([fahrt['zeile'], 0, 2])
                        extNotCalculated += 1
                        success = False

                if success:
                    fahrt['orig_kosten'] = (fahrt['ger_kosten']*dz) + fahrt['mautk']
                    orders.append(fahrt)

            extOrders[kz] = orders

def getZonePreise():
    edvSheet = edvBook.sheet_by_index(0)

    for i in range(2, 24):
        zone = edvSheet.cell(i, 0).value
        if isinstance(zone, float):
            zone = int(zone)
        zone = str(zone)
        allowedZones.append(zone)

        prices = [None]*3
        lastprice = 0;
        for y in range(0, 3):

            price = edvSheet.cell(i, 3-y).value
            if price != '':
                lastprice = price
            else:
                price = lastprice

            prices[2 - y] = price

        zonePrices[zone] = prices

    for i in range(27, 36):
        zone = edvSheet.cell(i, 0).value
        zone = str(zone)

        allowedZones.append(zone)

        price = float(edvSheet.cell(i, 1).value)
        prices = [price]*3

        zonePrices[zone] = prices

def getStundenPreise():
    edvSheet = edvBook.sheet_by_index(1)

    hourPrices["2A"] = float(edvSheet.cell(1, 0).value)
    hourPrices["3A"] = float(edvSheet.cell(1, 1).value)
    hourPrices["4A"] = float(edvSheet.cell(1, 2).value)
    hourPrices["Bagger"] = float(edvSheet.cell(1, 2).value)
    hourPrices["4A+"] = float(edvSheet.cell(1, 3).value)
    hourPrices["5A"] = float(edvSheet.cell(1, 4).value)
    hourPrices["HZ"] = float(edvSheet.cell(1, 4).value)
    hourPrices["Kran"] = float(edvSheet.cell(1, 5).value)

def calculateInterns():

    global KZOrdersCalc
    KZOrdersCalc = KZOrders

    global allIntCalcDZ

    for kz, orders in KZOrdersCalc.items():
        for fahrt in orders:
            fahrt['dz_abs'] = fahrt['ger_kosten'] * (dz - 1)
            allIntCalcDZ += fahrt['dz_abs']
            fahrt['summe'] = (fahrt['ger_kosten'] * dz) + fahrt['mautk']

def calculateSubcontractors():
    global extOrdersCalc
    global minus
    extOrdersCalc = extOrders

    #Liste bereinigen
    for kz, orders in extOrdersCalc.items():
        for fahrt in orders:

            #Sack ausrechnen
            doMind = True
            if fahrt['einheit'] == 'Sack':
                fahrt['anmerkungen'] = str(int(fahrt['menge'])) + " Säcke geladen. 40 Säcke pro Tonne laut Tarifblatt"
                fahrt['menge'] = fahrt['menge']/40
                fahrt['einheit'] = 'to'
                doMind = False

            #Mindermenge + Ausnahme aus EDV_Abrechnung.xls
            do = True
            for rule in Excemptions['KeineMindermenge']:
                isZero = len(rule['trigger'])
                for specifier in rule['trigger']:
                    if specifier[1] in fahrt[specifier[0]]:
                        isZero -= 1
                if isZero == 0:
                    do = False
                    break
            if do and fahrt['einheit'] == 'to':
                if fahrt['menge'] < mindermenge and doMind:
                    fahrt['anmerkungen'] = str(fahrt['menge']) + "to - Mindermenge - 12to gerechnet."
                    fahrt['menge'] = mindermenge

            dodz = True
            if fahrt['einheit'] == 'pau':
                dodz = False
            else:

                #if kz.lower() == "uuhope3":
                #    if fahrt['lfs_nr'] == 501714:
                #        print("hallo1")
                #    if fahrt['zone'] == 4 or fahrt['zone'] == '4':
                #        print("hallo1")
                #    print("heraussen")

                if "2Achs" in fahrt['geraet']:
                    z = 0
                    st = '2A'
                elif "3Achs" in fahrt['geraet']:
                    z = 0
                    st = '3A'
                elif "4Achs" in fahrt['geraet']:
                    z = 1
                    st = '4A'
                elif "5Achs" in fahrt['geraet']:
                    z = 2
                    st = '5A'
                else:
                    success = False

            #Tonnage Preise rechnen + Ausnahmen
            if fahrt['einheit'] == 'to':
                newPrice = False
                for rule in Excemptions['PreisProTonne']:
                    isZero = len(rule['trigger'])
                    for specifier in rule['trigger']:
                        if specifier[1] in fahrt[specifier[0]]:
                            isZero -= 1
                        if isZero == 0:
                            newPrice = float(rule['action'])
                            if "HOPE3" in fahrt["kz"]:
                                if fahrt['lfs_nr'] == 498692:
                                    print("hi")
                            break

                if not newPrice:
                    preis = zonePrices[fahrt['zone']][z]
                    fahrt['ger_kosten'] = float(fahrt['menge'])*preis
                elif newPrice:
                    fahrt['ger_kosten'] = float(fahrt['menge'])*newPrice

            #Stunden Preise rechnen + Ausnahmen
            if fahrt['einheit'] == 'std':
                newPrice = False
                for rule in Excemptions['PreisProStunde']:
                    isZero = len(rule['trigger'])
                    for specifier in rule['trigger']:
                        if specifier[1] in fahrt[specifier[0]]:
                            isZero -= 1
                        if isZero == 0:
                            newPrice = float(rule['action'])

                if not newPrice:
                    preis = hourPrices[st]
                else:
                    preis = newPrice

                if fahrt['stunden'] == '' or fahrt['stunden'] is None:
                    fahrt['ger_kosten'] = float(fahrt['menge'])*preis
                else:
                    fahrt['ger_kosten'] = float(fahrt['stunden'])*preis

            #Dieselzuschalg einrechnen + Ausnahme

            for rule in Excemptions['KeinTZ']:
                isZero = len(rule['trigger'])
                for specifier in rule['trigger']:
                    if specifier[1] in fahrt[specifier[0]]:
                        isZero -= 1
                    if isZero == 0:
                        dodz = False
                        break
            DZ = dz
            if dodz:
                DZ += 0.02
            else:
                DZ = 1

            fahrt['dz_abs'] = fahrt['ger_kosten'] * (DZ-1)

            do = True;
            for rule in Excemptions['KeineMaut']:
                isZero = len(rule['trigger'])
                for specifier in rule['trigger']:
                    if specifier[1] in fahrt[specifier[0]]:
                        isZero -= 1
                if isZero == 0:
                    do = False
                    break
            if not do and fahrt['mautk'] > 0:
                fahrt['mautk'] = 0
                fahrt['anmerkungen'] += "Maut wie vereinbart auf 0€ gesetzt"

            fahrt['summe'] = (fahrt['ger_kosten'] * DZ) + fahrt['mautk']

            #Machen wir ein Minusgschäft?
            if fahrt['summe'] > fahrt['orig_kosten']:
                minus += (fahrt['summe']-fahrt['orig_kosten'])
                fehler.append([fahrt['zeile'], 0, 3, fahrt])

def printKFZ(orders, extern):


    if extern:
        DZ = dz+0.02
    else:
        DZ = dz

    for kz, fahrten in orders.items():
        outBook = openpyxl.Workbook()
        outSheet = outBook.active

        try:
            outSheet.cell(1, 1).value = kz
            if extern:
                outSheet.cell(1, 2).value = 'Subfrächter Gutschrift'
            else:
                outSheet.cell(1, 2).value = 'Erlösblatt'

            if kz in extKZZuordnung:
                outSheet.cell(1, 4).value = extKZZuordnung[kz]

            outSheet.cell(3, 1).value = 'Gerät'
            outSheet.cell(3, 2).value = 'LFS-Datum'
            outSheet.cell(3, 3).value = 'LFS-Nr'
            outSheet.cell(3, 4).value = 'Baustelle'
            outSheet.cell(3, 5).value = 'Zone'
            outSheet.cell(3, 6).value = 'Einheit'
            outSheet.cell(3, 7).value = 'Menge'
            outSheet.cell(3, 8).value = 'Ger. Kosten'
            outSheet.cell(3, 9).value = 'DZ'
            outSheet.cell(3, 10).value = 'Mautkosten'
            outSheet.cell(3, 11).value = 'Summe'
            outSheet.cell(3, 12).value = 'Anmerkungen'

            ger_sum = 0
            maut_sum = 0
            dz_sum = 0

            i = 0

            for fahrt in fahrten:
                try:
                    ger_sum += float(fahrt['ger_kosten'])
                    maut_sum += float(fahrt['mautk'])
                    dz_sum += float(fahrt['dz_abs'])

                    outSheet.cell(i + 4, 1).value = fahrt['geraet']
                    outSheet.cell(i + 4, 2).value = fahrt['lfs_datum']
                    outSheet.cell(i + 4, 3).value = fahrt['lfs_nr']
                    outSheet.cell(i + 4, 4).value = fahrt['baustelle']
                    outSheet.cell(i + 4, 5).value = fahrt['zone']
                    outSheet.cell(i + 4, 6).value = fahrt['einheit']
                    outSheet.cell(i + 4, 7).value = fahrt['menge']
                    outSheet.cell(i + 4, 8).value = round(fahrt['ger_kosten'], 2)
                    outSheet.cell(i + 4, 9).value = round(float(fahrt['dz_abs']), 2)
                    outSheet.cell(i + 4, 10).value = fahrt['mautk']
                    outSheet.cell(i + 4, 11).value = round(fahrt['summe'], 2)
                    outSheet.cell(i + 4, 12).value = fahrt['anmerkungen']

                    i += 1
                except:
                    print("Fatal Error Writing to Output XLSX - Zeile" + fahrt['Zeile'] )
                    traceback.print_exc()

            ger_sum = round(ger_sum, 2)
            maut_sum = round(maut_sum, 2)
            dz_sum = round(dz_sum, 2)

            #Descriptors
            outSheet.cell(i + 6, 8).value = 'Ger. Kosten'
            outSheet.cell(i + 8, 8).value = 'Dieselzuschlag'
            outSheet.cell(i + 9, 8).value = 'Maut'
            outSheet.cell(i + 10, 8).value = 'Gesamtsumme'
            #Diesel
            outSheet.cell(i + 8, 10).value = str(round(((DZ-1)*100), 1)) + '%'
            #Values
            outSheet.cell(i + 6, 11).value = ger_sum
            outSheet.cell(i + 8, 11).value = dz_sum
            outSheet.cell(i + 9, 11).value = maut_sum
            outSheet.cell(i + 10, 11).value = ger_sum + dz_sum + maut_sum

            styleSheet(outSheet)

            #Speichern
            if extern:
                outBook.save(saveintern + "Subfrächter_" + date + "/" + kz + '.xlsx')
            else:
                outBook.save(saveintern + date + "/" + kz + '_i.xlsx')
        except:
            print("Fatal-Error - printKFZ() - KZ: " + kz)
            traceback.print_exc()

def styleSheet(sheet):

    none = openpyxl.styles.Side(style=None)
    thin = openpyxl.styles.Side(style='thin')
    dotted = openpyxl.styles.Side(style='dotted')
    underline = openpyxl.styles.borders.Border(
        left=none,
        right=none,
        top=dotted,
        bottom=thin,
    )

    overline = openpyxl.styles.borders.Border(
        left=none,
        right=none,
        top=thin,
        bottom=none,
    )

    # apply Corner Styles
    sheet.row_dimensions[1].height = 20
    sheet.cell(1, 1).font = openpyxl.styles.Font(bold=True)

    for i in range(4, sheet.max_row+1, 1):
        sheet.cell(i, 2).number_format = 'DD.MM.YYYY'
        sheet.cell(i, 8).number_format = u'"€" #,##0.00'
        sheet.cell(i, 9).number_format = u'"€" #,##0.00'
        sheet.cell(i, 10).number_format = u'"€" #,##0.00'
        sheet.cell(i, 11).number_format = u'"€" #,##0.00'

    # apply header styles
    for i in range(1, sheet.max_column + 1, 1):
        sheet.cell(3, i).border = underline
        sheet.cell(3, i).font = openpyxl.styles.Font(bold=True)

        # apply final sum styles
        for i in range(sheet.max_column - 4, sheet.max_column, 1):
            sheet.cell(sheet.max_row, i).border = underline
            sheet.cell(sheet.max_row, i).font = openpyxl.styles.Font(bold=True)

        # apply inter sum style
        for i in range(sheet.max_column - 4, sheet.max_column, 1):
            sheet.cell(sheet.max_row - 4, i).border = overline
            sheet.cell(sheet.max_row - 4, i).font = openpyxl.styles.Font(bold=True)


    widths = [
        ['A', 10],
        ['B', 11],
        ['C', 11],
        ['D', 15],
        ['E', 4],
        ['F', 6],
        ['G', 9],
        ['H', 10],
        ['I', 9],
        ['J', 7],
        ['K', 12],
        ['L', 15],
    ]

    for i in range(0, sheet.max_column, 1):
        sheet.column_dimensions[widths[i][0]].width = widths[i][1]

    sheet.page_setup.orientation = sheet.ORIENTATION_LANDSCAPE
    sheet.page_setup.paperSize = sheet.PAPERSIZE_A4

def readExcemptionos():
    edvSheet = edvBook.sheet_by_index(2)

    global mindermenge
    global sackTonne

    sackTonne = int(edvSheet.cell(0,1).value)
    mindermenge = int(edvSheet.cell(1,1).value)

    for r in range(6, edvSheet.nrows):
        name = edvSheet.cell(r, 0).value
        type = edvSheet.cell(r, 1).value
        action = edvSheet.cell(r, 2).value
        if type != '' and type is not None:

            thisRule = {'action': action, 'trigger': [], 'name': name}
            rule = edvSheet.cell(r, 0).value

            if type not in Excemptions:
                Excemptions[type] = []

            for rr in range(6, edvSheet.nrows):
                if rule == edvSheet.cell(rr, 4).value:
                    trigger = edvSheet.cell(rr, 5).value
                    value = edvSheet.cell(rr, 6).value
                    if isinstance(value, float):
                        value = str(int(value))
                    thisRule['trigger'].append([trigger, value])

            Excemptions[type].append(thisRule)

def loadConfig():
    global recipients
    global inFolder
    global outFolder
    global errFolder
    global backupFolder
    global AgMappingPath
    global TourMappingPath
    global sourcePath

    global edvbookPath
    global saveintern
    global abrFilePath
    global savespotPath
    global extbookPath

    with open('config.csv', 'r') as f:
        lines = f.readlines()
        print(lines)
        for x in range(len(lines)):
            lineElems = lines[x].split(';')

            if x == 0:
                edvbookPath = lineElems[1]

            if x == 1:
                saveintern = lineElems[1]

            if x == 2:
                abrFilePath = lineElems[1]

            if x == 3:
                savespotPath = lineElems[1]

            if x == 4:
                extbookPath = lineElems[1]

main()